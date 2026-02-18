"""
excel_export.py — Build watchlist results Excel workbook
Two sheets: Summary + All Trades
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── Colour palette ──
DARK_BG   = "0C1419"
HEADER_BG = "0A2A3A"
ACCENT    = "00E5B4"
ACCENT2   = "0099FF"
BULL      = "00C896"
BEAR      = "FF4D6D"
WARN      = "FF9A6C"
WHITE     = "FFFFFF"
LIGHT_BG  = "EAF7F4"
BLUE_BG   = "E8F4FF"
MUTED     = "5A7A8A"
ROW_ALT   = "F5FFFE"

thin = Side(style="thin", color="C0DDD8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr(ws, cell, value, bg=HEADER_BG, fg=ACCENT, bold=True, sz=10, wrap=False, center=True):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=wrap)
    c.border = BORDER


def _cell(ws, cell, value, bold=False, color="1A2E38", bg=None, center=False, fmt=None):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=9)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def build_excel(watchlist_results: list) -> bytes:
    """
    watchlist_results: list of dicts, one per ticker:
    {
      ticker, price, as_of,
      scan: { iv: {...}, fusion5: {...} },
      backtest: { iv: { stats, trades }, fusion5: { stats, trades } }
    }
    Returns xlsx as bytes.
    """
    wb = Workbook()

    # ════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    ws_sum.freeze_panes = "A3"

    # Title row
    ws_sum.merge_cells("A1:R1")
    t = ws_sum["A1"]
    t.value = "StockSense AI — Watchlist Backtest Summary"
    t.font  = Font(name="Arial", bold=True, color=ACCENT, size=14)
    t.fill  = PatternFill("solid", fgColor=DARK_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 32

    # Column headers — row 2
    headers = [
        ("A2","Ticker"),
        ("B2","Price"),
        ("C2","Scan — IV"),
        ("D2","Scan — Fusion5"),
        # IV backtest
        ("E2","IV Trades"),
        ("F2","IV Win%"),
        ("G2","IV Avg Ret%"),
        ("H2","IV Best%"),
        ("I2","IV Worst%"),
        ("J2","IV Max DD%"),
        ("K2","IV Profit Factor"),
        # Fusion5 backtest
        ("L2","F5 Trades"),
        ("M2","F5 Win%"),
        ("N2","F5 Avg Ret%"),
        ("O2","F5 Best%"),
        ("P2","F5 Worst%"),
        ("Q2","F5 Max DD%"),
        ("R2","F5 Profit Factor"),
    ]
    for cell, label in headers:
        col = cell[0] if len(cell) == 2 else cell[:2]
        bg  = HEADER_BG
        fg  = ACCENT if "IV" in label or "Ticker" in label or "Price" in label or "Scan" in label else ACCENT2
        _hdr(ws_sum, cell, label, bg=bg, fg=fg, sz=9, wrap=True)

    ws_sum.row_dimensions[2].height = 36

    # Column widths
    col_widths = [10,10,18,18,10,9,12,10,10,10,14,10,9,12,10,10,10,14]
    for i, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for row_i, res in enumerate(watchlist_results, start=3):
        ticker  = res.get("ticker", "")
        price   = res.get("price", "")
        scan    = res.get("scan", {})
        bt      = res.get("backtest", {})
        row_bg  = ROW_ALT if row_i % 2 == 0 else WHITE

        # Scan signals
        iv_scan = scan.get("iv", {}).get("status", "—")
        f5_scan = scan.get("fusion5", {}).get("status", "—")
        scan_iv_disp  = "✅ Entry!" if iv_scan == "ENTRY_TRIGGERED" else ("👁 Watching" if iv_scan == "WATCHING" else "— No Signal")
        scan_f5_disp  = "✅ Entry!" if f5_scan == "ENTRY_TRIGGERED" else ("👁 Watching" if f5_scan == "WATCHING" else "— No Signal")

        iv_s  = bt.get("iv",      {}).get("stats", {})
        f5_s  = bt.get("fusion5", {}).get("stats", {})

        row = row_i
        _cell(ws_sum, f"A{row}", ticker,   bold=True, color=DARK_BG, bg=row_bg)
        _cell(ws_sum, f"B{row}", price,    color=DARK_BG, bg=row_bg, fmt='"$"#,##0.00')

        # Scan status cells
        def scan_cell(cell, txt):
            c = ws_sum[cell]
            c.value = txt
            c.font  = Font(name="Arial", size=9, bold=True,
                          color=BULL if "Entry" in txt else (WARN if "Watch" in txt else MUTED))
            c.fill  = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

        scan_cell(f"C{row}", scan_iv_disp)
        scan_cell(f"D{row}", scan_f5_disp)

        def stat_cell(cell, val, is_pct=False, is_pf=False):
            if not val and val != 0:
                _cell(ws_sum, cell, "—", color=MUTED, bg=row_bg, center=True)
                return
            v = float(val)
            if is_pct:
                color = BULL if v > 0 else (BEAR if v < 0 else MUTED)
                _cell(ws_sum, cell, v/100, color=color, bg=row_bg, center=True, fmt='0.0%')
            elif is_pf:
                color = BULL if v >= 1.5 else (WARN if v >= 1 else BEAR)
                _cell(ws_sum, cell, v, color=color, bg=row_bg, center=True, fmt='0.00"x"')
            else:
                _cell(ws_sum, cell, v, bg=row_bg, center=True)

        # IV stats
        _cell(ws_sum, f"E{row}", iv_s.get("total","—"), bg=row_bg, center=True)
        stat_cell(f"F{row}",  iv_s.get("win_rate"),    is_pct=True)
        stat_cell(f"G{row}",  iv_s.get("avg_return"),  is_pct=True)
        stat_cell(f"H{row}",  iv_s.get("best_trade"),  is_pct=True)
        stat_cell(f"I{row}",  iv_s.get("worst_trade"), is_pct=True)
        stat_cell(f"J{row}",  iv_s.get("max_drawdown"),is_pct=True)
        stat_cell(f"K{row}",  iv_s.get("profit_factor"), is_pf=True)

        # Fusion5 stats
        _cell(ws_sum, f"L{row}", f5_s.get("total","—"), bg=row_bg, center=True)
        stat_cell(f"M{row}",  f5_s.get("win_rate"),    is_pct=True)
        stat_cell(f"N{row}",  f5_s.get("avg_return"),  is_pct=True)
        stat_cell(f"O{row}",  f5_s.get("best_trade"),  is_pct=True)
        stat_cell(f"P{row}",  f5_s.get("worst_trade"), is_pct=True)
        stat_cell(f"Q{row}",  f5_s.get("max_drawdown"),is_pct=True)
        stat_cell(f"R{row}",  f5_s.get("profit_factor"), is_pf=True)

        ws_sum.row_dimensions[row].height = 20

    # Totals row
    n_data = len(watchlist_results)
    if n_data > 0:
        tot_row = 3 + n_data
        ws_sum.merge_cells(f"A{tot_row}:D{tot_row}")
        _hdr(ws_sum, f"A{tot_row}", "WATCHLIST TOTALS", bg="0A2A3A", fg=ACCENT, sz=9)

        for col_letter, start_row in [
            ("E", 3), ("F", 3), ("G", 3), ("H", 3), ("I", 3),
            ("J", 3), ("K", 3), ("L", 3), ("M", 3), ("N", 3),
            ("O", 3), ("P", 3), ("Q", 3), ("R", 3)
        ]:
            end_r  = tot_row - 1
            c_addr = f"{col_letter}{tot_row}"
            if col_letter in ("E", "L"):   # trade count — sum
                ws_sum[c_addr] = f"=SUM({col_letter}3:{col_letter}{end_r})"
            else:                           # averages
                ws_sum[c_addr] = f"=IFERROR(AVERAGE({col_letter}3:{col_letter}{end_r}),\"—\")"
            ws_sum[c_addr].font      = Font(name="Arial", bold=True, color=ACCENT, size=9)
            ws_sum[c_addr].fill      = PatternFill("solid", fgColor=HEADER_BG)
            ws_sum[c_addr].alignment = Alignment(horizontal="center", vertical="center")
            ws_sum[c_addr].border    = BORDER
            if col_letter in ("F","G","H","I","J","M","N","O","P","Q"):
                ws_sum[c_addr].number_format = "0.0%"
            elif col_letter in ("K","R"):
                ws_sum[c_addr].number_format = '0.00"x"'
        ws_sum.row_dimensions[tot_row].height = 22

    # ════════════════════════════════════════
    # SHEET 2 — ALL TRADES
    # ════════════════════════════════════════
    ws_tr = wb.create_sheet("All Trades")
    ws_tr.sheet_view.showGridLines = False
    ws_tr.freeze_panes = "A3"

    # Title
    ws_tr.merge_cells("A1:L1")
    t2 = ws_tr["A1"]
    t2.value = "StockSense AI — All Trades Log (2-Year Backtest)"
    t2.font  = Font(name="Arial", bold=True, color=ACCENT2, size=14)
    t2.fill  = PatternFill("solid", fgColor=DARK_BG)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_tr.row_dimensions[1].height = 32

    tr_headers = [
        ("A2","Ticker"), ("B2","Strategy"), ("C2","Entry Date"),
        ("D2","Entry $"), ("E2","Stop Loss $"), ("F2","Target (2R) $"),
        ("G2","Exit Date"), ("H2","Exit $"), ("I2","Result"),
        ("J2","Return %"), ("K2","Duration (days)"), ("L2","Scan Signal"),
    ]
    for cell, label in tr_headers:
        fg = ACCENT if label in ("Ticker","Strategy","Entry Date","Entry $") else ACCENT2
        _hdr(ws_tr, cell, label, fg=fg, sz=9, wrap=True)
    ws_tr.row_dimensions[2].height = 32

    tr_col_widths = [10,14,14,12,12,14,14,12,12,12,16,14]
    for i, w in enumerate(tr_col_widths, 1):
        ws_tr.column_dimensions[get_column_letter(i)].width = w

    tr_row = 3
    for res in watchlist_results:
        ticker   = res.get("ticker", "")
        bt       = res.get("backtest", {})
        scan     = res.get("scan", {})
        iv_scan  = scan.get("iv",      {}).get("status", "")
        f5_scan  = scan.get("fusion5", {}).get("status", "")
        scan_lbl = []
        if iv_scan == "ENTRY_TRIGGERED":   scan_lbl.append("IV ✅")
        if f5_scan == "ENTRY_TRIGGERED":   scan_lbl.append("F5 ✅")
        scan_str = ", ".join(scan_lbl) if scan_lbl else "No signal"

        for strat_key in ("iv", "fusion5"):
            trades = bt.get(strat_key, {}).get("trades", [])
            for t in trades:
                row_bg = ROW_ALT if tr_row % 2 == 0 else WHITE
                ret    = t.get("return_pct")
                reason = t.get("exit_reason", "")
                ret_color = BULL if (ret or 0) > 0 else BEAR

                _cell(ws_tr, f"A{tr_row}", ticker,              bold=True, bg=row_bg)
                _cell(ws_tr, f"B{tr_row}", t.get("strategy",""),bg=row_bg)
                _cell(ws_tr, f"C{tr_row}", t.get("entry_date",""),bg=row_bg, center=True)
                _cell(ws_tr, f"D{tr_row}", t.get("entry_price"),bg=row_bg, center=True, fmt='"$"#,##0.00')
                _cell(ws_tr, f"E{tr_row}", t.get("sl"),          bg=row_bg, center=True, fmt='"$"#,##0.00', color=BEAR)
                _cell(ws_tr, f"F{tr_row}", t.get("target"),      bg=row_bg, center=True, fmt='"$"#,##0.00', color=BULL)
                _cell(ws_tr, f"G{tr_row}", t.get("exit_date",""),bg=row_bg, center=True)
                _cell(ws_tr, f"H{tr_row}", t.get("exit_price"),  bg=row_bg, center=True, fmt='"$"#,##0.00')

                # Result tag
                reason_display = {"TARGET":"✅ Target","TARGET1":"✅ T1 Hit","TARGET2":"✅ T2 Hit",
                                  "SL":"❌ SL Hit","TIMEOUT":"⏱ Timeout"}.get(reason, reason)
                reason_color   = BULL if "Target" in reason_display or "Hit" in reason_display and "SL" not in reason_display else (BEAR if "SL" in reason_display else MUTED)
                _cell(ws_tr, f"I{tr_row}", reason_display, bold=True, color=reason_color, bg=row_bg)

                ret_val = ret/100 if ret is not None else None
                _cell(ws_tr, f"J{tr_row}", ret_val, color=ret_color, bg=row_bg, center=True,
                      fmt='+0.0%;-0.0%;"-"')
                _cell(ws_tr, f"K{tr_row}", t.get("duration_days"), bg=row_bg, center=True)
                _cell(ws_tr, f"L{tr_row}", scan_str, bg=row_bg, color=BULL if "✅" in scan_str else MUTED)

                ws_tr.row_dimensions[tr_row].height = 18
                tr_row += 1

    # ── Save to bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
